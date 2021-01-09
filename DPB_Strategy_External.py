from common_functions_ext import n_decim, two_decim, dt_to_str, str_to_dt
from convert_to_ohlc_ext import TradingViewData
from update_data_ext import UpdateStockData
from create_batch import CreateBatch
from pandas_datareader.data import get_data_yahoo as ydata
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import timedelta
from bs4 import BeautifulSoup
from pathlib import Path
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import datetime
import re
import json
import requests
import copy
import os
import threading
import time
import openpyxl


def period_info(period):
    prd_count = period['count']
    prd_start = period['start']
    prd_end = period['end']
    prd_type = period['type']
    return prd_count, prd_start, prd_end, prd_type


def hg_lw_finder(potential_index, lst, find_hg=True):
    peak_or_trough = True
    for cl in lst:
        if cl > lst[potential_index] and find_hg:
            peak_or_trough = False
        elif cl < lst[potential_index] and not find_hg:
            peak_or_trough = False
    return peak_or_trough


def num_to_pct(num):
    return 1 - (num / 100)


class Stock:
    '''
    Contains methods needed to backtest Double Point Breakout Strategy'''

    def __init__(self, ticker, methods_args, freq="1D"):
        '''
        Parameters
            ticker   = ticker of stock
            freq  = interval of stock data
            '''

        self.ticker = ticker.upper()
        self.tf = freq
        self.ticker_tf = f"{self.ticker}_{freq}"

        args = methods_args['convert_data']
        self.convert_data(args['use_existing'])

        args = methods_args['get_ohlc_table']
        self.get_ohlc_table(args['start_dt'], args['end_dt'], args['use_yfin'],
                            args['same_folder'])

        # self.get_earnings()

        args = methods_args['get_periods']
        self.get_periods(args['pct_drop'], args['sec_drop'], args['only_rec'])

        args = methods_args['get_resistances']
        self.get_resistances(args['stop_pct'], args['take_pct'],
                             args['default_pct'], args['res_diff'],
                             args['hld_bef'], args['hld_aft'])

        args = methods_args['get_triggers']
        self.get_triggers(args['days_apart'], args['inv_days'],
                          args['hld_bef'], args['hld_aft'], args['hg_diff'])

        args = methods_args['get_trades']
        self.get_trades(args['aut_stp_pct'], args['sell_pct'],
                        args['sec_sell'], args['hld_days'])

        self.get_results()

        args = methods_args['export_pandas']
        self.export_pandas(args['export'])

        args = methods_args['export_excel']
        self.export_excel(args['export'])

    def convert_data(self, use_existing):
        if not use_existing:
            data_df = TradingViewData(self.ticker, self.tf, same_folder=True)
            data_df.latest_download()
            data_df.read_raw_csv()
            data_df.to_pickle()

    def get_ohlc_table(self, start_dt, end_dt, use_yfin, same_folder):
        '''
        Uses self.ticker or self.ticker_tf to select stocks
            Stores pandas dataframe of stock data it in self.ohlc_table

            start_dt = start date of stock data
            end_dt   = end date of stock data
            use_yfin = use pandas_datareader to get stock data from
                        yahoo finance, default opens data from downloaded
                        data from TradinView'''

        if not use_yfin:
            if not same_folder:
                file_dir = r"E:\Python\Stock_Data\1D\\"
                data_path = f"{file_dir}{self.ticker_tf}.pkl"
            else:
                file_dir = Path(Path.cwd())/self.tf
                data_path = file_dir/f'{self.ticker_tf}.pkl'

            data = pd.read_pickle(data_path)

            if not start_dt and not end_dt:
                self.ohlc_table = data

            elif start_dt and not end_dt:
                self.ohlc_table = data[start_dt:]

            elif not start_dt and end_dt:
                self.ohlc_table = data[: end_dt]

            else:
                self.ohlc_table = data[start_dt: end_dt]

        else:
            if start_dt == '' and end_dt == '':
                stock = ydata(self.ticker)
            elif start_dt == '':
                stock = ydata(self.ticker, end=end_dt)
            elif end_dt == '':
                stock = ydata(self.ticker, start=start_dt)
            else:
                stock = ydata(self.ticker, start=start_dt, end=end_dt)
            self.ohlc_table = stock[['Open', 'High', 'Low', 'Close']]

    def get_earnings(self):
        '''
        Gets stock historical earnings releases dates from scraping yahoo
        finance and stores them in lst_earnings_dt

            lst_earnings_dt value types = strings, format(YYYY-MM-DD)'''

        # TODO Fix this method
        self.lst_earnings_dt = []
        url_e_releases = '''https://ca.finance.yahoo.com/calendar/earnings?day=2020-10-11&symbol=MSFT'''
        response = requests.get(url_e_releases)
        soup = BeautifulSoup(response.text, 'html.parser')
        pattern = re.compile(r'\s--\sData\s--\s')
        script_data = soup.find('script', text=pattern).contents[0]
        start = script_data.find('context') - 2
        json_data = json.loads(script_data[start:-12])
        lst_earnings_dt = json_data['context']['dispatcher']['stores']['ScreenerResultsStore']['results']['rows']

        for earning_release in lst_earnings_dt:
            self.lst_earnings_dt.append(earning_release["startdatetime"][:10])

    def get_periods(self, pct_drop, sec_drop, only_rec):
        '''
        Finds periods that qualify and stores information
        about them in lst_prd

            pct_drop = minimum % for period to start
            sec_drop = minimum % for secondary period to start

            lst_prd = list of dictionaries
                     = {prd_count, prd_start, prd_end, prd_type}

            Keys   = Values                        = Datatype

            prd_count = Labels periods starting from 1   = Integer
            prd_start = Date when period starts       = String
            prd_end   = Date when period ends         = String
            prd_type  = If period recovered(R) or not(D) = String'''

        def reset_prd_var(dip=False):
            '''
            Resets variables used to find periods

                dip = True if period didn't recover and fell into
                    secondary period

                Resets = self.cur_hgst_p, self.prd_start,
                        self.cur_lwst_p and sec_hgst_p

                dip = False = Resets variables to default values

                dip = True  = Sets self.sec_hgst_p as
                            self.cur_hgst_p, the price to recover
                            from new period.
                         = Sets self.prd_start as the days date
                         = Resets rest of variables to default'''

            if not dip:
                self.cur_hgst_p = 0
                self.prd_start = None
            else:
                self.cur_hgst_p = self.sec_hgst_p
                self.prd_start = dt_to_str(dt)
            self.cur_lwst_p, self.sec_hgst_p = 100000000, 0

        def test_pre_prd(years=10):
            '''
            Sees if data goes back {years} years from a date'''

            test_dt = str(int(self.prd_start[:4]) - years)
            return self.ohlc_table[test_dt]

        def lst_prd_append(dip=False):
            '''
            Appends period data to self.lst_prd'''

            dict_prd = {"count": prd_count, "start": self.prd_start,
                        "end": dt_to_str(dt)}
            if not dip:
                dict_prd['type'] = "R"
            else:
                dict_prd['type'] = "D"

            self.lst_prd.append(dict_prd)

        def dict_prd_stats_add():
            '''
            Adds tuple with high price before period and lowest price
            during period to self.dict_prd_stats dictionnary'''

            self.dict_prd_stats[prd_count] = (self.cur_hgst_p, self.cur_lwst_p)

        pct_drop = num_to_pct(pct_drop)
        sec_drop = num_to_pct(sec_drop)

        reset_prd_var()
        prd_count = 1
        self.lst_prd, self.dict_prd_stats = [], {}

        for day in self.ohlc_table.itertuples():
            (dt, op, hg, lw, cl) = day

            if not self.prd_start:
                if cl > self.cur_hgst_p:
                    self.cur_hgst_p = cl

                if cl < (self.cur_hgst_p * pct_drop):
                    self.prd_start = dt_to_str(dt)
            else:
                if cl > self.cur_hgst_p:
                    try:
                        test_pre_prd()

                    except KeyError:
                        reset_prd_var()
                        continue

                    lst_prd_append()
                    dict_prd_stats_add()
                    reset_prd_var()
                    prd_count += 1
                    continue

                if cl < self.cur_lwst_p:
                    self.cur_lwst_p = cl

                if not only_rec:

                    if cl > (self.cur_lwst_p * 1.5) and cl > self.sec_hgst_p:
                        self.sec_hgst_p = cl

                    if self.sec_hgst_p and cl < (self.sec_hgst_p * sec_drop):
                        try:
                            test_pre_prd()

                        except KeyError:
                            reset_prd_var(dip=True)
                            continue

                        lst_prd_append(dip=True)
                        dict_prd_stats_add()
                        reset_prd_var(dip=True)
                        prd_count += 1

    def get_resistances(self, stop_pct, take_pct, default_pct,
                        res_diff, hld_bef, hld_aft):
        '''
        Finds resistance lines, their stop loss and take profit
        for each period
            Stores data in self.dict_res and self.dict_res_stats

            stop_pct = stop loss percentage below resistance
            take_pct = take profit percentage below resistance
            default_pct = maximum resistance % below previous resistance
            res_diff = minimum pct difference from lower resistance
                        to resistance above it
            hld_bef  = minimum days before a peak for it to count
            hlf_aft  = minimum days after a peak for it to count

            self.dict_res  = {n: (prd_dates, prd_type, lst_res_prices)}
            Keys        = Values                        = Datatype

            prd_dates   = Period start date and end date   = String
                            = Separated by comma (1 string)
            prd_type    = If period recovered(R) or not(D) = String
            lst_res_prices = List of resistance prices     = List[Floats]

            self.dict_res_stats = {n: lst_res_stats}
            lst_res_stats  = (res_p, res_type, res_stop, res_take)
            Keys  = Values                                  = Datatype

            res_p = Resistance price                        = Float
            res_type = If set using peak and troughs(A) or not(D) = String
            res_stop = Resistance stop loss price              = Float
            res_take = Resistance take profit                  = Float'''

        def get_stop(res_p):
            '''
            Returns stop loss below resistance price
                res_p = Resistance price
                pct   = Number percent below resistance price
            '''
            pct_below = num_to_pct(stop_pct)
            return n_decim(res_p * pct_below)

        def get_take(res_p):
            '''
            Returns take profit below resistance price
                res_p = Resistance price
                pct   = Number percent below resistance price
            '''
            pct_below = num_to_pct(take_pct)
            return n_decim(res_p * pct_below)

        def default_res(res_p, round=False):
            '''
            Returns default resistance price ({default_pct}% below higher
            resistance)
                res_p = Resistance price
                pct   = Furthest percent from higher to next resistance price
                round = Whether to use n_decim to round answer
            '''
            pct_lower = num_to_pct(default_pct)
            if not round:
                return res_p * pct_lower
            if round:
                return n_decim(res_p * pct_lower)

        res_diff = num_to_pct(res_diff)
        total_hld = hld_bef + hld_aft + 1
        peak_i = hld_bef
        self.dict_res, self.dict_hg_lw, self.dict_res_stats = {}, {}, {}
        for period in self.lst_prd:
            prev_days_cl, lst_highs_lows = [], []
            (prd_count, prd_start, prd_end, prd_type) = period_info(period)

            month_day = prd_start[4:-2] + "01"
            ten_y = str(int(prd_start[:4])-10) + month_day

            init_resist = n_decim(self.dict_prd_stats[prd_count][0])
            init_stop = get_stop(init_resist)
            init_take = get_take(init_resist)
            lst_res_prices = [(init_resist, 'A')]
            lst_res_stats = [(init_resist, 'A', init_stop, init_take)]

            for day in self.ohlc_table[ten_y:prd_start].itertuples():
                (dt, op, hg, lw, cl) = day

                if len(prev_days_cl) > total_hld:
                    del prev_days_cl[0]

                    if hg_lw_finder(peak_i, prev_days_cl) or\
                            hg_lw_finder(peak_i, prev_days_cl, find_hg=False):
                        lst_highs_lows.append(prev_days_cl[peak_i])

                prev_days_cl.append(cl)
            lst_highs_lows = sorted(lst_highs_lows, reverse=True)

            for (res_p, _) in lst_res_prices:
                nxt_res_p = None

                for hg_lw in lst_highs_lows:
                    if hg_lw <= res_p * res_diff:

                        if hg_lw > default_res(res_p):
                            nxt_res_p = n_decim(hg_lw)

                    if hg_lw < default_res(res_p):
                        break

                res_type = "A"
                if not nxt_res_p:
                    nxt_res_p = default_res(res_p, round=True)
                    res_type = "D"
                    lst_res_prices.append((nxt_res_p, res_type))

                else:
                    lst_res_prices.append((nxt_res_p, res_type))

                nxt_stp = get_stop(nxt_res_p)
                nxt_take = get_take(nxt_res_p)
                lst_res_stats.append((nxt_res_p, res_type, nxt_stp, nxt_take))

                if nxt_res_p < self.dict_prd_stats[prd_count][1]:
                    break

            prd_dates = f"{prd_start}, {prd_end}"
            self.dict_hg_lw[prd_count] = lst_highs_lows
            self.dict_res[prd_count] = prd_dates, prd_type, lst_res_prices
            self.dict_res_stats[prd_count] = lst_res_stats

    def get_triggers(self, days_apart, inv_days, hld_bef,
                     hld_aft, hg_diff):
        '''
        Finds entry trigger prices for each period
            Stores trigger datas in self.dict_triggers

            Finds all peaks and if their prices are less than {hg_diff}
            percent apart, are more than {days_apart} and no closes above
            pot_p1 in between highs, store this trigger data in dictionary.

            days_apart = minimum days between peaks to count as trigger
            inv_days   = number of days each triggers are valid
            hld_bef = minimum days before peak for it to be valid
            hld_aft = minimum days after peak for it to be valid
            hg_diff = maximum percentage difference between first peak
                        to second peak be considered as an entry trigger

            self.dict_triggers = {n: lst_trg}
            lst_trg         = (pot_dt1, pot_p1, pot_dt2, pot_inv_dt1)

            Variables   = Values                 = Datatype

            pot_dt1  = date of first peak     = String
            pot_p1   = entry trigger price    = Float
            pot_dt2  = date of second peak    = Datetime
            pot_inv_dt1 = 6 months after first peak = Datetime
                     = Or date when price closes higher than pot_p1
        '''
        def lst_trg_add():
            '''
            Appends trigger datas to lst_trg
            '''
            return dt_to_str(pot_dt1), pot_p1, pot_dt2, pot_inv_dt1

        self.dict_triggers = {}
        inv_dt = datetime.timedelta(inv_days)
        six_months = datetime.timedelta(180)

        hg_diff = hg_diff / 100
        total_hld = hld_bef + hld_aft + 1
        peak_i = hld_bef

        for period in self.lst_prd:
            prev_days_cl, prev_days_dt, lst_pot_trg, lst_trg = [], [], [], []
            (prd_count, prd_start, prd_end, prd_type) = period_info(period)

            pre_period_start = dt_to_str(str_to_dt(prd_start) - six_months)
            selected_period = self.ohlc_table[pre_period_start:prd_end]

            for day in selected_period.itertuples():
                (dt, op, hg, lw, cl) = day

                if len(prev_days_cl) > total_hld:
                    del prev_days_cl[0]
                    del prev_days_dt[0]

                    if hg_lw_finder(peak_i, prev_days_cl):
                        hg_dt = prev_days_dt[peak_i]
                        hg_cl = prev_days_cl[peak_i]
                        lst_pot_trg.append([hg_dt, hg_cl, hg_dt + inv_dt])

                prev_days_cl.append(cl)
                prev_days_dt.append(dt)

                for pot_trigger in lst_pot_trg:
                    (_, pot_trg_p, pot_inv_dt) = pot_trigger

                    if cl > pot_trg_p and dt < pot_inv_dt:
                        pot_trigger[2] = dt

            for index, pot_trg1 in enumerate(lst_pot_trg.copy()):
                (pot_dt1, pot_p1, pot_inv_dt1) = pot_trg1

                for pot_trg2 in lst_pot_trg[index+1:].copy():
                    (pot_dt2, pot_p2, _) = pot_trg2
                    trg_pct_diff = (pot_p1 - pot_p2) / pot_p1

                    if pot_inv_dt1 >= pot_dt2 and np.busday_count(dt_to_str(
                            pot_dt1), dt_to_str(pot_dt2)) > days_apart and\
                            trg_pct_diff <= hg_diff and trg_pct_diff >= 0:

                        if len(lst_trg) > 0:
                            added_to_list = False

                            for index2, (trg_dt, *_) in enumerate(lst_trg):

                                if pot_dt1 < str_to_dt(trg_dt):
                                    pot_dt1 = dt_to_str(pot_dt1)
                                    lst_trg.insert(index2, lst_trg_add())
                                    added_to_list = True
                                    break

                            if not added_to_list:
                                lst_trg.append(lst_trg_add())

                        else:
                            lst_trg.append(lst_trg_add())
                        break

            self.dict_triggers[prd_count] = sorted(lst_trg)

    def get_trades(self, aut_stp_pct, sell_pct, sec_sell, hld_days):
        '''
        Finds all trades for each period
            Stores trade datas in self.dict_trades
            Stores returns in dollars in arr_rsult_p
            Stores returns in percent in arr_rsult_pct

            aut_stp_pct = stop loss % below entry price
            sell_pct = minimum % retracement from take high to trigger sell
            sec_sell = minimum % retracement from above next resistance
                           to trigger sell
            hld_days = days after trigger to enter if entry still valid

            self.dict_trades = {n: lst_trades}
            lst_trades    = (result_p, result_pct, ent_dt, ent_p, exit_dt,
                                 exit_p)
            Variables  = Values               = Datatype

            result_p   = Trade return in dollars = Float
            result_pct = Trade return in percent = Float
            ent_dt  = Entry date           = String
            ent_p   = Entry price          = Float
            exit_dt = Exit date            = String
            exit_p  = Exit price           = Float

            arr_rsult_p   = Array[Float]
            arr_rsult_pct = Array[Float]

        '''
        def reset_trading_var():
            '''
            Resets variables used in this method to default values
            '''
            self.ent_trggerd, self.in_trade = False, False
            self.ent_dt, self.sell_trggerd = None, False
            self.brk_tke_hg, self.aut_stp, self.ent_p = 0, 0, 0
            self.stp_loss, self.pre_ent_stp, self.hgst_brk_res_p = 0, 0, 0
            self.hld_days_lft = 0

        def add_trades(exit_p, use_prev_dt=False):
            '''
            Appends appropriate trade datas in arr_rsult_p,
            arr_rsult_pct and lst_trades
                Also calculates trade return in dollars and percentages
            '''
            self.rsult_p = n_decim(exit_p - self.ent_p)
            self.rsult_pct = n_decim((exit_p - self.ent_p) / self.ent_p * 100)
            self.arr_rsult_p = np.append(self.arr_rsult_p, self.rsult_p)
            self.arr_rsult_pct = np.append(self.arr_rsult_pct, self.rsult_pct)
            if not use_prev_dt:
                exit_dt = sdt
            else:
                exit_dt = prev_day_dt
            lst_trades.append((self.rsult_p, self.rsult_pct, self.ent_dt,
                               n_decim(self.ent_p), exit_dt, n_decim(exit_p)))

        def pct_change(p1, p2):
            '''
            Returns percent difference from one price(p1) to another(p2)
            '''
            return 1 - ((p1 - p2) / p1)

        def qualify_trigger(initial=True):
            lwr_res_stop, lwr_res_p = 0, 0

            for res_stats in sorted(self.dict_res_stats[prd_count]):
                (res_p, _, res_stop, res_take) = res_stats

                if res_p > cl:

                    if cl > self.pre_ent_stp and res_p - cl > cl - lwr_res_p:
                        self.pre_ent_stp = lwr_res_stop

                        if initial:
                            self.ent_trggerd = True
                            self.hld_days_lft = hld_days

                    elif not initial:
                        reset_trading_var()

                    break

                else:
                    lwr_res_stop = res_stop
                    lwr_res_p = res_p

        aut_stp_pct = num_to_pct(aut_stp_pct)
        sell_pct = num_to_pct(sell_pct)
        sec_sell = num_to_pct(sec_sell)
        self.hld_days = hld_days

        self.dict_trades = {}
        dict_triggers_copy = copy.deepcopy(self.dict_triggers)
        self.arr_rsult_p, self.arr_rsult_pct = np.array([]), np.array([])

        for period in self.lst_prd:
            lst_trades = []
            prev_day_dt, prev_day_cl = None, 0
            (prd_count, prd_start, prd_end, prd_type) = period_info(period)
            reset_trading_var()

            six_months = datetime.timedelta(180)
            pre_period_start = dt_to_str(str_to_dt(prd_start) - six_months)

            for day in self.ohlc_table[pre_period_start:prd_end].itertuples():
                (dt, op, hg, lw, cl) = day
                sdt = dt_to_str(dt)
                enter_new_trade = not self.in_trade and self.ent_trggerd

                if enter_new_trade and dt >= str_to_dt(prd_start):

                    if self.hld_days_lft == 0:

                        if op > self.pre_ent_stp:
                            self.ent_dt = sdt
                            self.ent_p = op
                            self.in_trade = True
                            self.aut_stp = n_decim(op * aut_stp_pct)
                        else:
                            reset_trading_var()
                    else:
                        if cl > op and cl > prev_day_cl:
                            qualify_trigger(initial=False)
                            self.hld_days_lft -= 1
                        else:
                            reset_trading_var()
                else:
                    self.ent_trggerd = False
                    self.pre_ent_stp = 0

                if self.sell_trggerd:
                    add_trades(op)
                    reset_trading_var()

                '''if self.in_trade and sdt in self.lst_earnings_dt:
                    if np.busday_count(self.ent_dt, sdt) > 4:
                        add_trades(prev_day_cl, use_prev_dt=True)

                    reset_trading_var()'''

                if self.in_trade:
                    higher_res_p, higher_res_take = 0, 0

                    for (res_p, *_) in self.dict_res_stats[prd_count]:

                        if cl > res_p:

                            if res_p > self.hgst_brk_res_p:
                                self.hgst_brk_res_p = res_p
                                self.brk_tke_hg = 0
                            break

                    for res_stats in self.dict_res_stats[prd_count]:
                        (res_p, _, res_stop, res_take) = res_stats
                        same_zone = res_p == self.hgst_brk_res_p

                        if res_stop < op and op < res_p and same_zone:

                            if hg > higher_res_take and pct_change(hg, cl) < sell_pct or\
                                    hg > higher_res_p and pct_change(hg, cl) < sec_sell:

                                if cl < higher_res_p:
                                    self.sell_trggerd = True

                            self.stp_loss = res_stop
                            break

                        elif res_p < op:

                            if hg > higher_res_p:

                                if cl < higher_res_p:

                                    if pct_change(hg, cl) < sec_sell:

                                        self.sell_trggerd = True
                                        self.brk_tke_hg = 0

                            if hg > higher_res_take:

                                if cl < higher_res_p:

                                    if hg > self.brk_tke_hg:
                                        self.brk_tke_hg = hg

                                    if pct_change(hg, cl) < sell_pct:

                                        self.sell_trggerd = True
                                        self.brk_tke_hg = 0

                            if self.brk_tke_hg:

                                if pct_change(self.brk_tke_hg, cl) < sell_pct:

                                    if cl < higher_res_p:
                                        self.sell_trggerd = True
                                        self.brk_tke_hg = 0

                            self.stp_loss = res_stop
                            break

                        higher_res_p = res_p
                        higher_res_take = res_take

                    if self.stp_loss > self.aut_stp:

                        if lw < self.stp_loss:

                            if op > self.stp_loss:
                                add_trades(self.stp_loss)
                            else:
                                add_trades(op)

                            reset_trading_var()
                    else:

                        if self.in_trade and lw < self.aut_stp:

                            if op >= self.aut_stp:
                                add_trades(self.aut_stp)
                            else:
                                add_trades(op)

                            reset_trading_var()

                prev_day_dt = sdt
                prev_day_cl = cl

                for trigger in dict_triggers_copy[prd_count].copy():
                    (trg_dt, trg_p, trg_start_dt, trg_end_dt) = trigger

                    if cl > trg_p and dt > trg_start_dt or dt > trg_end_dt:
                        dict_triggers_copy[prd_count].remove(trigger)

                    if self.in_trade or self.ent_trggerd:
                        continue

                    if cl > trg_p and dt == trg_end_dt:
                        qualify_trigger()

            previous_day_loss = False

            breakout_day = True
            after_prd_end = str_to_dt(prd_end) + six_months

            if self.in_trade:

                for day in self.ohlc_table[prd_end:after_prd_end].itertuples():
                    (dt, op, hg, lw, cl) = day
                    hgst_stp = n_decim(self.dict_res_stats[prd_count][0][2])
                    self.stp_loss = hgst_stp

                    if cl < self.stp_loss and not breakout_day:
                        add_trades(self.stp_loss)
                        break

                    if previous_day_loss:

                        if cl < op:
                            add_trades(cl)
                            break

                        if cl > op:
                            previous_day_loss = False

                    if cl < op:
                        previous_day_loss = True

                    breakout_day = False

            self.dict_trades[prd_count] = lst_trades

    def get_results(self):
        '''
        Prints results_summary of stock's historical performance
            Stats                = Definition

            Total #Trades        = Total number of trades
            Total #Wins          = Total number of wins
            Total #Losses        = Total number of losses
            Win Percentage       = Win percentage (wins/total trades*100)
            Loss Win Ratio       = Loss win ratio (losses / wins)
            Average Win Percentage  = Average percent return per win
            Average Loss Percentage = Average percent return per loss
            Total Percentage Return = Total return in percentages
            Total Price Return   = Total return in dollars
            Profitability        = Total profitability
            Regular Win Percentage  = Average percent return per regular win
            Regular Loss Percentage = Average percent return per regular loss
            Regular Percentage   = Total regular moves percent return
            Regular Profitability   = Regular moves profitability

            Regular moves = Wins < 15% and losses < 5%
            Profitability = (Average win percentage - average loss percentage
                              * loss win ratio) / Average win percentage
        '''
        def fmt_str(strg, value, pct=True, dol=False):
            '''
            Formats string for asthetic purposes
                pct = Whether to return in percentage
                dol = Whether to return in dollars
            '''
            if dol:
                return strg.ljust(35, ".") + str(value).rjust(8) + " $" + "\n"

            elif pct:
                return strg.ljust(35, ".") + str(value).rjust(8) + " %" + "\n"

            else:
                return strg.ljust(35, ".") + str(value).rjust(8) + "\n"

        def profit_pct(regular=False):
            '''
            Calculates profitability
                regular = Whether to only calculate profitability for
                           regular moves
            '''
            try:
                if not regular:
                    in_decim = avg_win_pct + (avg_loss_pct * lss_win_ratio)
                    return two_decim(in_decim / avg_win_pct * 100)

                else:
                    in_decim = reg_win_avg_pct + \
                        (reg_loss_avg_pct * lss_win_ratio)
                    return two_decim(in_decim / reg_win_avg_pct * 100)

            except TypeError:
                return None

        def arr_avg(array):
            '''
            Returns properly rounded average of arrays values
            '''
            if array.sum() == 0:
                return np.nan
            return two_decim(np.average(array))

        def arr_sum(array):
            '''
            Returns properly rounded sum of arrays values
            '''
            if two_decim(array.sum()) != 0:
                return two_decim(array.sum())
            else:
                return np.nan

        def arr_count(array):
            '''
            Returns number of values in arrays
            '''
            if np.count_nonzero(array) > 0:
                return np.count_nonzero(array)
            else:
                return np.nan

        def get_summary(df, stats=None, results=True):
            if results:
                number_trades = df['#Trades'].sum()
                number_wins = df['#Wins'].sum()
                number_losses = df['#Losses'].sum()
                lss_win_ratio = number_losses / number_wins
                num_reg_wins = stats['# Reg Wins'].sum()
                num_reg_losses = stats['# Reg Losses'].sum()
                avg_win_pct = stats['W % Return'].sum() / number_wins
                avg_loss_pct = stats['L % Return'].sum() / number_losses
                reg_win_avg_pct = stats['Reg W % Return'].sum() / num_reg_wins
                reg_loss_avg_pct = stats['Reg L % Return'].sum(
                ) / num_reg_losses

                dataframe = pd.DataFrame({
                    '#Hold Days': np.nan,
                    '#Trades': number_trades,
                    '#Wins': number_wins,
                    '#Losses': number_losses,
                    'Win %': number_wins / number_trades * 100,
                    'L/W Ratio': lss_win_ratio,
                    'Av. Win %': avg_win_pct,
                    'Av. Loss %': avg_loss_pct,
                    '% Return': df['% Return'].sum(),
                    '$ Return': df['$ Return'].sum(),
                    'Profit. %': (avg_win_pct + (avg_loss_pct * lss_win_ratio)) / avg_win_pct * 100,
                    'Reg Win %': reg_win_avg_pct,
                    'Reg Loss %': reg_loss_avg_pct,
                    'Reg % Return': df['Reg % Return'].sum(),
                    'Reg Profit. %': (reg_win_avg_pct + (reg_loss_avg_pct * lss_win_ratio)) / reg_win_avg_pct * 100
                }, index=['Summary']).round(2)

            else:
                dataframe = pd.DataFrame({
                    '#Hold Days': np.nan,
                    'W % Return': df['W % Return'].sum(),
                    'L % Return': df['L % Return'].sum(),
                    '# Reg Wins': df['# Reg Wins'].sum(),
                    '# Reg Losses': df['# Reg Losses'].sum(),
                    'Reg W % Return': df['Reg W % Return'].sum(),
                    'Reg L % Return': df['Reg L % Return'].sum()},
                    index=['Summary']).round(2)

            return dataframe

        if not self.dict_trades:
            return

        trade_number = arr_count(self.arr_rsult_pct)
        pct_return = arr_sum(self.arr_rsult_pct)
        arr_wins = self.arr_rsult_pct[self.arr_rsult_pct > 0]
        avg_win_pct = arr_avg(arr_wins)
        wins_number = arr_count(arr_wins)
        arr_reg_wins = arr_wins[arr_wins < 15]
        reg_win_avg_pct = arr_avg(arr_reg_wins)

        arr_losses = self.arr_rsult_pct[self.arr_rsult_pct < 0]
        avg_loss_pct = arr_avg(arr_losses)
        losses_number = arr_count(arr_losses)
        arr_reg_losses = arr_losses[arr_losses > -5]
        reg_loss_avg_pct = arr_avg(arr_reg_losses)

        price_return = arr_sum(self.arr_rsult_p)
        arr_reg_moves = self.arr_rsult_pct[self.arr_rsult_pct < 15]
        arr_reg_moves = arr_reg_moves[arr_reg_moves > -5]
        reg_moves_pct_return = arr_sum(arr_reg_moves)

        try:
            win_percentage = two_decim(wins_number/trade_number*100)
            lss_win_ratio = two_decim(losses_number/wins_number)
            profitability = profit_pct()
            reg_profitability = profit_pct(regular=True)
        except ZeroDivisionError:
            win_percentage, lss_win_ratio, profitability = np.nan, np.nan, np.nan
            reg_profitability = np.nan
        except TypeError:
            win_percentage, lss_win_ratio, profitability = np.nan, np.nan, np.nan
            reg_profitability = np.nan

        results_df = pd.DataFrame({
            '#Hold Days': self.hld_days,
            '#Trades': trade_number,
            '#Wins': wins_number,
            '#Losses': losses_number,
            'Win %': win_percentage,
            'L/W Ratio': lss_win_ratio,
            'Av. Win %': avg_win_pct,
            'Av. Loss %': avg_loss_pct,
            '% Return': pct_return,
            '$ Return': price_return,
            'Profit. %': profitability,
            'Reg Win %': reg_win_avg_pct,
            'Reg Loss %': reg_loss_avg_pct,
            'Reg % Return': reg_moves_pct_return,
            'Reg Profit. %': reg_profitability
        }, index=[self.ticker])

        stats_df = pd.DataFrame({
            '#Hold Days': self.hld_days,
            'W % Return': arr_sum(arr_wins),
            'L % Return': arr_sum(arr_losses),
            '# Reg Wins': arr_count(arr_reg_wins),
            '# Reg Losses': arr_count(arr_reg_losses),
            'Reg W % Return': arr_sum(arr_reg_wins),
            'Reg L % Return': arr_sum(arr_reg_losses)
        }, index=[self.ticker])

        lst_results.append(results_df)
        lst_stats.append(stats_df)

    def export_pandas(self, export):
        def get_summary(df, stats=None, results=True):
            if results:
                number_trades = df['#Trades'].sum()
                number_wins = df['#Wins'].sum()
                number_losses = df['#Losses'].sum()
                lss_win_ratio = number_losses / number_wins
                num_reg_wins = stats['# Reg Wins'].sum()
                num_reg_losses = stats['# Reg Losses'].sum()
                avg_win_pct = stats['W % Return'].sum() / number_wins
                avg_loss_pct = stats['L % Return'].sum() / number_losses
                reg_win_avg_pct = stats['Reg W % Return'].sum() / num_reg_wins
                reg_loss_avg_pct = stats['Reg L % Return'].sum(
                ) / num_reg_losses

                dataframe = pd.DataFrame({
                    '#Hold Days': np.nan,
                    '#Trades': number_trades,
                    '#Wins': number_wins,
                    '#Losses': number_losses,
                    'Win %': number_wins / number_trades * 100,
                    'L/W Ratio': lss_win_ratio,
                    'Av. Win %': avg_win_pct,
                    'Av. Loss %': avg_loss_pct,
                    '% Return': df['% Return'].sum(),
                    '$ Return': df['$ Return'].sum(),
                    'Profit. %': (avg_win_pct + (avg_loss_pct * lss_win_ratio)) / avg_win_pct * 100,
                    'Reg Win %': reg_win_avg_pct,
                    'Reg Loss %': reg_loss_avg_pct,
                    'Reg % Return': df['Reg % Return'].sum(),
                    'Reg Profit. %': (reg_win_avg_pct + (reg_loss_avg_pct * lss_win_ratio)) / reg_win_avg_pct * 100
                }, index=['Summary']).round(2)

            else:
                dataframe = pd.DataFrame({
                    '#Hold Days': np.nan,
                    'W % Return': df['W % Return'].sum(),
                    'L % Return': df['L % Return'].sum(),
                    '# Reg Wins': df['# Reg Wins'].sum(),
                    '# Reg Losses': df['# Reg Losses'].sum(),
                    'Reg W % Return': df['Reg W % Return'].sum(),
                    'Reg L % Return': df['Reg L % Return'].sum()},
                    index=[self.ticker]).round(2)

            return dataframe

        def call_summary(results, stats):
            results_summary = get_summary(results, stats=stats)
            stats_summary = get_summary(stats, results=False)
            return results_summary, stats_summary

        def replace_data():
            all_results.iloc[real_i] = result_df
            all_stats.iloc[real_i] = stats_df

            return all_results[:-1], all_stats[:-1]

        def append_data():
            new_results = pd.concat([all_results[:-1], result_df])
            new_stats = pd.concat([all_stats[:-1], stats_df])
            return new_results, new_stats

        def sort_df(results, stats):
            results.sort_index(inplace=True)
            stats.sort_index(inplace=True)

        def append_summary(results_df, stats_df):
            all_results = pd.concat([results_df, results_summary])
            all_stats = pd.concat([stats_df, stats_summary])
            return all_results, all_stats

        def pkl_save():
            all_results.to_pickle(results_filepath)
            all_stats.to_pickle(stats_filepath)

        if not export:
            return

        os.makedirs('Results', exist_ok=True)
        os.makedirs('Stats', exist_ok=True)
        results_filepath = Path(Path.cwd())/'Results'/'1d_hld.pkl'
        stats_filepath = Path(Path.cwd())/'Stats'/'1d_hld.pkl'

        try:
            all_results = pd.read_pickle(results_filepath)
            all_stats = pd.read_pickle(stats_filepath)

            for result_df, stats_df in zip(lst_results, lst_stats):
                ticker = result_df.index[0]
                hld_days = result_df['#Hold Days'][0]

                try:
                    df_part = all_results[all_results['#Hold Days'] == int(
                        hld_days)][ticker:ticker]['#Hold Days']

                    if result_df['#Hold Days'].equals(df_part):
                        pot_i = np.flatnonzero(
                            all_results['#Hold Days'] == hld_days)
                        real_i = []
                        for i in pot_i.copy():
                            if all_results.iloc[i].name == ticker:
                                real_i.append(i)
                                break

                        if real_i:
                            new_results, new_stats = replace_data()
                        else:
                            new_results, new_stats = append_data()
                    else:
                        new_results, new_stats = append_data()

                    sort_df(new_results, new_stats)
                    results_summary, stats_summary = call_summary(
                        new_results, new_stats)
                    all_results, all_stats = append_summary(
                        new_results, new_stats)
                except KeyError:
                    new_results, new_stats = append_data()
                    sort_df(new_results, new_stats)
                    results_summary, stats_summary = call_summary(
                        new_results, new_stats)
                    all_results, all_stats = append_summary(
                        new_results, new_stats)

            pkl_save()
        except FileNotFoundError:

            all_results = pd.DataFrame(lst_results[0])
            all_stats = pd.DataFrame(lst_stats[0])

            for result_df in lst_results[1:]:
                all_results = pd.concat([all_results, result_df])

            for stat in lst_stats[1:]:
                all_stats = pd.concat([all_stats, stat])

            results_summary, stats_summary = call_summary(
                all_results, all_stats)

            sort_df(all_results, all_stats)

            all_results, all_stats = append_summary(all_results, all_stats)

            pkl_save()

    def export_excel(self, export):

        if not export:
            return

        def export_data(df):

            dataframe = pd.read_pickle(f'{df.title()}\\1d_hld.pkl')
            xlsx_filepath = Path(Path.cwd())/df.title()/'1d_hld.xlsx'
            f'{df.title()}\\1d_hld.xlsx'
            dataframe.to_excel(xlsx_filepath)

            data_wb = openpyxl.load_workbook(xlsx_filepath)
            data_sheet = data_wb['Sheet1']

            data_sheet.freeze_panes = 'B2'
            data_sheet.column_dimensions['A'].width = 12

            centered = Alignment(horizontal='center')
            ticker_font = Font(size=16)
            data_font = Font(size=13)

            for row_num in range(2, data_sheet.max_row):
                data_sheet.row_dimensions[row_num].height = 20
                data_cell = data_sheet['A' + str(row_num)]
                data_cell.font = ticker_font

            for column_num in range(2, data_sheet.max_column + 1):
                column_letter = get_column_letter(column_num)
                data_sheet.column_dimensions[column_letter].width = 15

                for row_num in range(2, data_sheet.max_row + 1):
                    data_cell = data_sheet[column_letter + str(row_num)]
                    data_cell.alignment = centered
                    data_cell.font = data_font
                    if row_num % 2 == 0:
                        grey_fill = PatternFill("solid", fgColor='00C0C0C0')
                        data_cell.fill = grey_fill

            data_wb.save(xlsx_filepath)

        export_data('results')
        export_data('stats')

    def visualise(self):
        def get_dates(date):
            date1 = dt_to_str(str_to_dt(date) - timedelta(line_length))
            date2 = dt_to_str(str_to_dt(date) + timedelta(line_length))
            return date1, date2

        self.dict_plots, self.trades_plots = {}, {}
        for period in self.lst_prd:
            (prd_count, prd_start, prd_end, prd_type) = period_info(period)
            closes, dates = [], []

            for (date, *_, close) in self.ohlc_table[prd_start:prd_end].itertuples():
                closes.append(close)
                dates.append(date)

            plt.subplots(figsize=(20, 8))
            plt.plot(dates, closes, color='black',
                     linestyle='dotted', linewidth='.5')

            for (price, *_) in self.dict_res_stats[prd_count]:
                graph = plt.hlines(price, xmin=dates[0],
                                   xmax=dates[-1], linestyle='dotted')

            for trade_data in self.dict_trades[prd_count]:
                (*_, entry_dt, entry_p, exit_dt, exit_p) = trade_data
                line_length = len(closes) * 0.01
                entry_d1, entry_d2 = get_dates(entry_dt)
                exit_d1, exit_d2 = get_dates(exit_dt)

                plt.hlines(entry_p, xmin=entry_d1,
                           xmax=entry_d2, color='green')
                graph = plt.hlines(exit_p, xmin=exit_d1,
                                   xmax=exit_d2, color='red')
                self.trades_plots[prd_count] = graph
            self.dict_plots[prd_count] = graph
        plt.show(block=True)


ticker_regex = re.compile(r'([A-Z]+)_1D\.pkl')
default_args = {
    'convert_data': {'use_existing': False},
    'get_ohlc_table': {'start_dt': '', 'end_dt': '', 'use_yfin': False,
                       'same_folder': True},
    'get_periods': {'pct_drop': 20, 'sec_drop': 30, 'only_rec': False},
    'get_resistances': {'stop_pct': .5, 'take_pct': 1.5, 'default_pct': 9,
                        'res_diff': 7.5, 'hld_bef': 3, 'hld_aft': 3},
    'get_triggers': {'days_apart': 7, 'inv_days': 180, 'hld_bef': 3,
                     'hld_aft': 7, 'hg_diff': 2},
    'get_trades': {'aut_stp_pct': 2, 'sell_pct': 1, 'sec_sell': .5,
                   'hld_days': 0.0},
    'export_pandas': {'export': False},
    'export_excel': {'export': False}
}

lst_results, lst_stats, stock_threads = [], [], []
lst_tickers = []
os.makedirs('1D', exist_ok=True)
UpdateStockData('1D').update()

for filename in os.listdir(Path(Path.cwd())/'1D'):
    if filename.endswith('_1D.pkl'):
        ticker_element = ticker_regex.search(filename)
        ticker = ticker_element.group(1)
        lst_tickers.append(ticker)

ten_pct = round(len(lst_tickers) / 10, 1)
to_print = []
pct_progress = 10
for multiple in range(1, 10):
    to_print.append(ten_pct * multiple)

for index, ticker in enumerate(lst_tickers):

    for var in [0.0, 1.0, 2.0]:
        updated_args = copy.deepcopy(default_args)
        updated_args['get_trades']['hld_days'] = var
        updated_args['get_ohlc_table']['start_dt'] = '1990'
        if index + 1 == len(lst_tickers) and var == 2.0:
            updated_args['export_pandas']['export'] = True
            updated_args['export_excel']['export'] = True

        stock_data = threading.Thread(
            target=Stock, args=[ticker, updated_args])
        stock_threads.append(stock_data)
        stock_data.start()
        while threading.active_count() > 20:
            time.sleep(1)

    if to_print:
        if index > to_print[0]:
            print(f'{pct_progress}%')
            pct_progress += 10
            del to_print[0]

for stock_thread in stock_threads:
    stock_thread.join()
CreateBatch('Results')
CreateBatch('Stats')

print('100%\n')
print('Done')
